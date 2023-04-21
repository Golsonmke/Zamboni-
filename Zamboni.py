#Zund America 
#Zamboni v1.0 (Repairing card helper)
#Authored by Gary Olson
#4.20.23
#
#
#Program to fill out repair card, print the pdf repairing card and write data to excel spreadsheet.

import tkinter as tk
from tkinter import filedialog
import openpyxl
from openpyxl import Workbook
from fpdf import FPDF

class RepairCard:
    def __init__(self, root):
        self.root = root
        self.root.title("Repair Card")
        self.root.geometry("800x800")
        
        self.label_service_partner = tk.Label(root, text="Service partner:")
        self.label_service_partner.pack()
        self.entry_service_partner = tk.Entry(root)
        self.entry_service_partner.pack()
        
        self.label_date = tk.Label(root, text="Ship Date:")
        self.label_date.pack()
        self.entry_date = tk.Entry(root)
        self.entry_date.pack()
        
        self.label_customer = tk.Label(root, text="Customer:")
        self.label_customer.pack()
        self.entry_customer = tk.Entry(root)
        self.entry_customer.pack()

        self.label_customer = tk.Label(root, text="Article number:")
        self.label_customer.pack()
        self.entry_customer = tk.Entry(root)
        self.entry_customer.pack()

        self.label_customer = tk.Label(root, text="Article name:")
        self.label_customer.pack()
        self.entry_customer = tk.Entry(root)
        self.entry_customer.pack()

        self.label_customer = tk.Label(root, text="Serial number of article:")
        self.label_customer.pack()
        self.entry_customer = tk.Entry(root)
        self.entry_customer.pack()

        self.label_customer = tk.Label(root, text="Replaced by serial number:")
        self.label_customer.pack()
        self.entry_customer = tk.Entry(root)
        self.entry_customer.pack()

        self.label_customer = tk.Label(root, text="Serial number of machine:")
        self.label_customer.pack()
        self.entry_customer = tk.Entry(root)
        self.entry_customer.pack()

        self.label_customer = tk.Label(root, text="Case number")
        self.label_customer.pack()
        self.entry_customer = tk.Entry(root)
        self.entry_customer.pack()

        self.label_customer = tk.Label(root, text="Error description/remarks:")
        self.label_customer.pack()
        self.entry_customer = tk.Entry(root)
        self.entry_customer.pack()
        
        self.button_print = tk.Button(root, text="Print Repair Card", command=self.print_card)
        self.button_print.pack()
        
        self.button_export = tk.Button(root, text="Export Data to Excel", command=self.export_data)
        self.button_export.pack()
        
    def print_card(self):
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)
        pdf.cell(200, 10, txt="Repair Card", ln=1, align="C")
        pdf.cell(200, 10, txt=f"Name: {self.entry_name.get()}", ln=1, align="L")
        pdf.cell(200, 10, txt=f"Phone: {self.entry_phone.get()}", ln=1, align="L")
        pdf.cell(200, 10, txt=f"Email: {self.entry_email.get()}", ln=1, align="L")
        pdf.output("RepairCard.pdf")
        
    def export_data(self):
        filename = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Spreadsheet", "*.xlsx")])
        if filename:
            wb = Workbook()
            ws = wb.active
            ws.title = "Repair Card Data"
            ws['A1'] = "Name"
            ws['B1'] = "Phone"
            ws['C1'] = "Email"
            ws['A2'] = self.entry_name.get()
            ws['B2'] = self.entry_phone.get()
            ws['C2'] = self.entry_email.get()
            wb.save(filename)
            
root = tk.Tk()
app = RepairCard(root)
root.mainloop()
